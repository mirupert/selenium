#reconfig attempt!
# import modules/libraries
import glob #Michael
import time
from openpyxl import Workbook
import math
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
print ('Welcome to Favso!')
print ('formerly the Lightspeed Data Entry App')
# source file input
file_list = []
file_number = 1
print ('    \nList of files:')
for file in glob.glob(r'*.xlsx'):
    if 'newnumbers' in file:
        continue
    if 'Gloria' or 'gloria' in file :
        file_list.append(file)
        print ('    ', file_number, '--', file)
        file_number = file_number + 1
while True:
    if len(file_list) == 1:
        selected_file_number = 0
        print ('  One option...')
        break
    selected_file_number = input('Enter file number: ')
    try:
        selected_file_number = int(selected_file_number)-1
        file_list[selected_file_number]
        break
    except:
        print('Retry with a diffferent number')
        continue
print ('  Do you want to run with file: ', file_list[selected_file_number], '?')
print ('\n  (To ONLY create new number file enter "NN")')
yesno = input('  Please type Yes, No, or NN:  ')
if yesno in ['Y','y','yes','Yes','YES', 'NN']:
    source_xlsx = file_list[selected_file_number]
    source = source_xlsx.replace('.xlsx', '')
    new_file_name=source
    wb = load_workbook(source + '.xlsx')
else:
    while True:
        source = input('\nEnter source file name (no extension): ')
        new_file_name=source
        if source == 'quit' :
            quit()
        try:
            wb = load_workbook(source + '.xlsx')
            break
        except:
            print('File does not exist try again')
ws = wb.active

#ensure gloria report is formatted properly
if ws.cell(row=1, column=4).value != 'Phone':
    print ('Gloria report is wrong format\nMake sure Phone numbers are in D column')
    input('Press any key to quit')
    quit()
if ws.cell(row=1, column=5).value != 'Subtotal':
    print ('Gloria report is wrong format\nMake sure Subtotals are in E column')
    input('Press any key to quit')
    quit()

# counter variables
row_counter = 1
Completed_number_tracker = 0
number_successes = 0
new_ss_tracker = 1
recheck_tracker = 1
# Backup for new number sheet will be deleted eventually 1/3

# Creating and formating newnumbers workbook for manual followup
new_wbook = Workbook()
new_ws = new_wbook.active
new_ws.column_dimensions['A'].width = 15
new_ws.column_dimensions['B'].width = 6
new_ws.column_dimensions['C'].width = 36
new_ws['A1'] = ('Enter on ReUp:')
new_ws['B1'] = ('Points')



# Ensure spreadsheets can save
try:
    wb.save(source + '.xlsx')
    new_wbook.save(new_file_name+'newnumbers.xlsx')
except:
    print ('Could not access files')
    print ('Make sure spreadsheets are closed and try again')
    input('Press any key to quit')
    quit()

if yesno != 'NN':
    # login to lightspeed
    driver = webdriver.Chrome('chromedriver')
    try:
        driver.get('https://loyalty.lightspeedapp.com')
        time.sleep(0.5)
        uname = driver.find_element_by_id('subdomain')
    except:
        print ('Application failed: Check internet connection')
        driver.quit()
        input('Press any key to quit')
        quit()
    uname.send_keys('tealargo')
    time.sleep(0.5)
    email = driver.find_element_by_id('email')
    email.send_keys('kristinmiller67+scanner@gmail.com')
    time.sleep(0.5)
    password = driver.find_element_by_name('password')
    password.send_keys('Lightspeed1', Keys.RETURN)
    time.sleep(1) #2seconds -> 1 8/16
    customersbutton = driver.find_element_by_id('cat-user')
    customersbutton.click()
    time.sleep(1) #2seconds -> 1 8/16

    # find phone numbers and order cost
    for cell in ws.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True) :
        for phone_number in cell :
            row_counter=row_counter+ 1
            order_cost = ws.cell(row=row_counter, column=5).value
            order_cost = math.floor(order_cost)
            consent_type = ws.cell(row=row_counter, column=11).value

    # processing tracker
        if row_counter % 25 == 0:
            print(row_counter, ' rows completed')


    # enter phone number
        if order_cost == 0 :
            Completed_number_tracker = Completed_number_tracker + 1
            continue

    # consent check
        if consent_type == ('NO_CONSENT'):
            continue
        add_points_link = driver.find_element_by_xpath('//*[@id="cat-user"]/ul/li[2]/a')
    # 8/13 added sleep here due to errors
        time.sleep(2)
        try:
            add_points_link.click()
        except:
            print('add points delay row: ', row_counter)
            time.sleep(4)
            try:
                add_points_link.click()
                print('Retry success')
            except:
                print('Retry fail\nnumber will be found on the new number report')
                continue
        phone = driver.find_element_by_id('userName')
        phone.send_keys(phone_number)
        time.sleep(2)
        autocomp = driver.find_element_by_id('ui-id-1')
        try :
            autocomp.click()
            number_successes = number_successes + 1
        except :
            phone.clear()
            continue
        time.sleep(0.5)
        points = driver.find_element_by_id('points')
        points.clear()
        points.send_keys(order_cost)
        time.sleep(0.5)
        assign_points = driver.find_element_by_xpath('//*[@id="creditForm"]/div[5]/div/button')
        assign_points.click()
        time.sleep(2)
        try:
            submit_points = driver.find_element_by_xpath('/html/body/div[5]/div/div/div[3]/div/div/button[1]')
            submit_points.click()
        except:
            print ('Initial submission failed')
            time.sleep(5)
            try:
                submit_points = driver.find_element_by_xpath('/html/body/div[5]/div/div/div[3]/div/div/button[1]')
                submit_points.click()
            except:
                continue

    # Once points are assigned order_cost becomes 0
        ws.cell(row=row_counter, column=5).value = 0
        ws.cell(row=row_counter, column=6).value = 'DONE'
        try:
            wb.save(source + '.xlsx')
        except:
            wait_for_ok = input('Ok to continue:')
            wb.save(source + '.xlsx')
        time.sleep(0.5)
    driver.quit()
    print ('\nCustomers skipped (previously completed): ', Completed_number_tracker)

#find non-zero subtotals, reconfig 1/26/22
for subtotals in ws.iter_rows(min_row=2, min_col=5, max_col=5, values_only=True) :
    for subtotal_value in subtotals :
        recheck_tracker = recheck_tracker + 1
        missed_phone_number = ws.cell(row=recheck_tracker, column=4).value
        if subtotal_value >= 1 :
            if ws.cell(row=recheck_tracker, column=11).value == ('NO_CONSENT'):
                print ('no_consent found')
                continue
            new_ss_tracker = new_ss_tracker + 1
            new_ws[('A'+(str(new_ss_tracker)))] = missed_phone_number
            subtotal_value = math.floor(subtotal_value)
            new_ws[('B'+(str(new_ss_tracker)))] = subtotal_value
total_ws_successes = number_successes + Completed_number_tracker
new_ws['C1'] = ('Completed through row ' + str(row_counter-1))
new_ws['C2'] = ('Customers that recieved points: ' + str(total_ws_successes))
print ('\nRows completed: ', str(row_counter-1))
print ('\nNumber of customers who earned points: ', number_successes)
new_wbook.save(new_file_name+'newnumbers.xlsx')
print ('You can close the window now')
time.sleep(20)
quit()
